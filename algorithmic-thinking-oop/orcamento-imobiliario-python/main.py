from dataclasses import dataclass
from abc import ABC, abstractmethod
import csv
import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

CONTRACT_FEE = 2000.00
MAX_INSTALLMENTS = 5


# ----------------- Utilidades -----------------
def format_brl(value: float) -> str:
    s = f"{value:,.2f}"
    s = s.replace(",", "X").replace(".", ",").replace("X", ".")
    return f"R$ {s}"


def leia_int(msg, valido=lambda x: True, erro="Valor inválido. Digite novamente."):
    while True:
        try:
            v = int(input(msg))
            if not valido(v):
                print(erro)
                continue
            return v
        except ValueError:
            print("Entrada inválida. Digite um número inteiro.")


def leia_opcao(msg, opcoes_validas):
    while True:
        v = input(msg).strip()
        if v in opcoes_validas:
            return v
        print(f"Opção inválida. Escolha entre: {', '.join(opcoes_validas)}.")


def ensure_base_name(name: str) -> str:
    """
    Garante um nome-base sem extensão.
    Ex.: 'orcamento.csv' -> 'orcamento'
    """
    name = name.strip()
    if not name:
        return "parcelas_orcamento"
    return Path(name).stem


def get_output_dir() -> Path:
    """
    Retorna a pasta 'exemplos' dentro do projeto.
    """
    base_dir = Path(__file__).resolve().parent
    output_dir = base_dir / "exemplos"
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir


# ----------------- Modelos -----------------
class Imovel(ABC):
    @abstractmethod
    def calcular_aluguel(self) -> float:
        pass


@dataclass
class Apartamento(Imovel):
    quartos: int = 1
    vagas: int = 0
    tem_criancas: bool = True

    def calcular_aluguel(self) -> float:
        v = 700.0
        if self.quartos == 2:
            v += 200.0
        if self.vagas > 0:
            v += 300.0
        if not self.tem_criancas:
            v *= 0.95
        return round(v, 2)


@dataclass
class Casa(Imovel):
    quartos: int = 1
    vagas: int = 0

    def calcular_aluguel(self) -> float:
        v = 900.0
        if self.quartos == 2:
            v += 250.0
        if self.vagas > 0:
            v += 300.0
        return round(v, 2)


@dataclass
class Estudio(Imovel):
    vagas: int = 0

    def calcular_aluguel(self) -> float:
        v = 1200.0
        if self.vagas == 0:
            return round(v, 2)
        if self.vagas <= 2:
            return round(v + 250.0, 2)
        return round(v + 250.0 + (self.vagas - 2) * 60.0, 2)


@dataclass
class Orcamento:
    imovel: Imovel
    parcelas_contrato: int

    def validar(self):
        if not (1 <= self.parcelas_contrato <= MAX_INSTALLMENTS):
            raise ValueError(f"Parcelas devem estar entre 1 e {MAX_INSTALLMENTS}.")

    def aluguel_mensal(self) -> float:
        return round(self.imovel.calcular_aluguel(), 2)

    def parcela_contrato(self) -> float:
        return round(CONTRACT_FEE / self.parcelas_contrato, 2)

    def gerar_parcelas(self):
        aluguel = self.aluguel_mensal()
        parcela = self.parcela_contrato()
        linhas = []

        for mes in range(1, 13):
            pc = parcela if mes <= self.parcelas_contrato else 0.0
            linhas.append({
                "mes": mes,
                "aluguel": round(aluguel, 2),
                "parcela_contrato": round(pc, 2),
                "total": round(aluguel + pc, 2),
            })

        return linhas


# ----------------- Saída / Resumo -----------------
def print_resumo(orc: Orcamento, linhas):
    aluguel = orc.aluguel_mensal()
    parcela = orc.parcela_contrato()

    print("\n--- RESUMO DO ORÇAMENTO ---")
    print(f"Aluguel mensal: {format_brl(aluguel)}")
    print(f"Taxa de contrato: {format_brl(CONTRACT_FEE)} em {orc.parcelas_contrato}x de {format_brl(parcela)}")
    print(f"Total do 1º mês: {format_brl(aluguel + parcela)}")

    if orc.parcelas_contrato < 12:
        print(f"Total após quitar o contrato: {format_brl(aluguel)}")

    print("\nPrévia (3 primeiros meses):")
    for l in linhas[:3]:
        parcela_txt = format_brl(l["parcela_contrato"]) if l["parcela_contrato"] else "—"
        print(
            f"Mês {l['mes']:>2}: "
            f"aluguel={format_brl(l['aluguel'])}, "
            f"parcela={parcela_txt}, "
            f"total={format_brl(l['total'])}"
        )


# ----------------- Geração CSV -----------------
def salvar_csv(filepath: Path, linhas):
    campos = ["Mês", "Aluguel", "Parcela do Contrato", "Total"]

    with open(filepath, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(campos)

        for l in linhas:
            writer.writerow([
                l["mes"],
                format_brl(l["aluguel"]),
                format_brl(l["parcela_contrato"]),
                format_brl(l["total"]),
            ])


# ----------------- Geração Excel -----------------
def salvar_excel(filepath: Path, linhas):
    wb = Workbook()
    ws = wb.active
    ws.title = "Orçamento"

    headers = ["Mês", "Aluguel", "Parcela do Contrato", "Total"]
    ws.append(headers)

    for l in linhas:
        ws.append([
            l["mes"],
            l["aluguel"],
            l["parcela_contrato"],
            l["total"],
        ])

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows(min_row=2, min_col=2, max_col=4):
        for cell in row:
            cell.number_format = 'R$ #,##0.00'

    for row in ws.iter_rows(min_row=2):
        row[0].alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)

        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_length:
                max_length = len(value)

        ws.column_dimensions[col_letter].width = max_length + 3

    wb.save(filepath)


# ----------------- Função principal -----------------
def main():
    print("=== Orçamento Imobiliário - CSV e Excel ===\n")

    print("Tipo de imóvel:")
    print("1) Apartamento")
    print("2) Casa")
    print("3) Estúdio")
    tipo = leia_opcao("Escolha [1-3]: ", {"1", "2", "3"})

    if tipo == "1":
        quartos = leia_int(
            "Número de quartos (1 ou 2): ",
            valido=lambda q: q in (1, 2),
            erro="Somente 1 ou 2."
        )
        vagas = leia_int(
            "Vagas de garagem: ",
            valido=lambda v: v >= 0,
            erro="Não pode ser negativo."
        )
        tem_criancas = input("Há crianças? [s/n]: ").strip().lower() in {"s", "sim"}
        imovel = Apartamento(quartos=quartos, vagas=vagas, tem_criancas=tem_criancas)

    elif tipo == "2":
        quartos = leia_int(
            "Número de quartos (1 ou 2): ",
            valido=lambda q: q in (1, 2),
            erro="Somente 1 ou 2."
        )
        vagas = leia_int(
            "Vagas de garagem: ",
            valido=lambda v: v >= 0,
            erro="Não pode ser negativo."
        )
        imovel = Casa(quartos=quartos, vagas=vagas)

    else:
        vagas = leia_int(
            "Vagas de estacionamento: ",
            valido=lambda v: v >= 0,
            erro="Não pode ser negativo."
        )
        imovel = Estudio(vagas=vagas)

    parcelas = leia_int(
        f"Nº de parcelas do contrato [1..{MAX_INSTALLMENTS}]: ",
        valido=lambda p: 1 <= p <= MAX_INSTALLMENTS,
        erro=f"Parcelas devem estar entre 1 e {MAX_INSTALLMENTS}."
    )

    orc = Orcamento(imovel=imovel, parcelas_contrato=parcelas)
    orc.validar()

    output_dir = get_output_dir()

    nome_base = input("Nome base do arquivo (ENTER = parcelas_orcamento): ")
    nome_base = ensure_base_name(nome_base or "parcelas_orcamento")

    csv_path = output_dir / f"{nome_base}.csv"
    xlsx_path = output_dir / f"{nome_base}.xlsx"

    linhas = orc.gerar_parcelas()

    try:
        salvar_csv(csv_path, linhas)
        salvar_excel(xlsx_path, linhas)

        print_resumo(orc, linhas)
        print(f"\n✅ CSV salvo em: {csv_path}")
        print(f"✅ Excel salvo em: {xlsx_path}")
        print("Geração concluída com sucesso!")

    except OSError as e:
        print(f"\n❌ Erro ao salvar os arquivos: {e}")
    except Exception as e:
        print(f"\n❌ Erro inesperado: {e}")


if __name__ == "__main__":
    main()